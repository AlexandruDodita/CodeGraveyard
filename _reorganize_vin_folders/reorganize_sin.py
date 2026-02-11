r"""
SIN Folder Reorganizer v3
=========================
Copies C:\SIN leasing vehicle folders into C:\SIN_Changed, organized by VIN.
Original files are NEVER modified or deleted.

Features:
  - Dry-run by default, --execute to apply
  - Reads PDF contents to discover VINs (PyMuPDF, all pages, threaded)
  - Collision-safe: identical files skipped, different files renamed with _1, _2
  - Generates centralized inventory Excel (one row per VIN, document categories)
  - Streaming .jsonl log for crash-safety

Usage:
  python reorganize_sin.py                          # dry run
  python reorganize_sin.py --execute                # copy to SIN_Changed
  python reorganize_sin.py --execute --no-pdf       # skip PDF scanning
  python reorganize_sin.py --workers 4              # parallel PDF scanning
  python reorganize_sin.py --range-start 5 --range-end 8
  python reorganize_sin.py --rename-files           # standardize PDF filenames
  python reorganize_sin.py --root "D:\SIN"          # custom source root
  python reorganize_sin.py --output "D:\SIN_Changed"  # custom output root
"""

import gc
import os
import re
import sys
import json
import time
import shutil
import hashlib
import datetime
import subprocess
from pathlib import Path
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Optional

IS_WINDOWS = sys.platform == "win32"

_RETRY_ATTEMPTS = 5
_RETRY_BASE_DELAY = 0.1  # seconds, doubles each attempt
MAX_CROSS_COPY_VINS = 100  # PDFs with more VINs than this skip normal cross-copy

# ── Windows path helpers ─────────────────────────────────────────────────────

def _long(p) -> str:
    s = str(p)
    if IS_WINDOWS and not s.startswith("\\\\?\\"):
        s = os.path.abspath(s)
        s = "\\\\?\\" + s
    return s


def _exists(p) -> bool:
    return os.path.exists(_long(p))


try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    class tqdm:
        def __init__(self, iterable=None, total=None, **kw):
            self._iterable = iterable
            self._total = total
        def __iter__(self):
            return iter(self._iterable)
        def update(self, n=1): pass
        def close(self): pass
        def set_postfix_str(self, *a, **kw): pass
        @staticmethod
        def write(s): print(s)

try:
    import fitz
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pytesseract
    from PIL import Image
    import io
    HAS_OCR = True
except ImportError:
    HAS_OCR = False

# ── Configuration ────────────────────────────────────────────────────────────

SIN_ROOT = Path(r"C:\SIN")
OUTPUT_ROOT = Path(r"C:\SIN_Changed")

VIN_PATTERN = re.compile(r'(?<![A-Z0-9])([A-Z0-9]{17})(?![A-Z0-9])')
FL_PATTERN = re.compile(
    r'^FL\s*-\s*.+?\s*-\s*([A-Z0-9]{17}).*\.pdf$', re.IGNORECASE
)
SERIEC_PATTERN = re.compile(r'^seriec_([A-Z0-9]{17})_', re.IGNORECASE)

# Partition merging: "SINDICALIZARE FOO - Part 1" → "SINDICALIZARE FOO"
_PART_SUFFIX = re.compile(r'\s*-\s*Part\s*\d+\s*$', re.IGNORECASE)


def merge_partition_name(name: str) -> str:
    """Strip ' - Part N' suffix to merge related partitions into one output folder."""
    return _PART_SUFFIX.sub('', name).rstrip()

# ── VIN helpers ──────────────────────────────────────────────────────────────

def is_valid_vin(s: str) -> bool:
    if len(s) != 17: return False
    return any(c.isalpha() for c in s) and any(c.isdigit() for c in s)


def is_vin(name: str) -> bool:
    name = name.strip()
    return bool(re.match(r'^[A-Z0-9]{17}$', name)) and is_valid_vin(name)


def extract_all_vins(fn: str) -> list:
    return [v for v in VIN_PATTERN.findall(fn) if is_valid_vin(v)]


def extract_vin_from_filename(fn: str) -> Optional[str]:
    m = FL_PATTERN.match(fn)
    if m: return m.group(1)
    m = SERIEC_PATTERN.match(fn)
    if m: return m.group(1)
    m = re.match(r'^([A-Z0-9]{17})[\s_\-]', fn)
    if m and is_valid_vin(m.group(1)): return m.group(1)
    return None


# ── PDF VIN extraction (cached, threaded) ────────────────────────────────────

_pdf_cache: dict = {}          # path_str -> set of VINs
_pdf_content_cats: dict = {}   # path_str -> set of critical category names
_pdf_stats = {"scanned": 0, "failed": 0, "vins_found": 0}

# ── Persistent OCR cache (survives across runs) ─────────────────────────────
# Keyed by path_str → {size, mtime, vins, cats, ocr_used}
# If file size/mtime changed since last scan, entry is stale → rescan.
_ocr_disk_cache: dict = {}
_OCR_CACHE_FILE = "ocr_cache.json"


def _file_fingerprint(path_str: str) -> tuple:
    """Return (size, mtime_ns) for a file, or (0, 0) on error."""
    try:
        st = os.stat(_long(path_str))
        return (st.st_size, int(st.st_mtime_ns))
    except OSError:
        return (0, 0)


def load_ocr_cache(output_root: Path):
    """Load persistent OCR cache from disk."""
    global _ocr_disk_cache
    cache_path = output_root / _OCR_CACHE_FILE
    if cache_path.exists():
        try:
            with open(str(cache_path), 'r', encoding='utf-8') as f:
                _ocr_disk_cache = json.load(f)
            print(f"  OCR cache: loaded {len(_ocr_disk_cache)} entries from {cache_path.name}",
                  file=sys.stderr)
        except Exception as exc:
            print(f"  WARNING: Could not load OCR cache: {exc}", file=sys.stderr)
            _ocr_disk_cache = {}
    else:
        _ocr_disk_cache = {}


def save_ocr_cache(output_root: Path):
    """Save persistent OCR cache to disk."""
    cache_path = output_root / _OCR_CACHE_FILE
    try:
        with open(str(cache_path), 'w', encoding='utf-8') as f:
            json.dump(_ocr_disk_cache, f, ensure_ascii=False)
        print(f"  OCR cache: saved {len(_ocr_disk_cache)} entries to {cache_path.name}",
              file=sys.stderr)
    except Exception as exc:
        print(f"  WARNING: Could not save OCR cache: {exc}", file=sys.stderr)


# ── Persistent rename map (original filenames survive across runs) ───────────
_RENAME_MAP_FILE = "rename_map.json"


def load_rename_map(output_root: Path) -> dict:
    """Load rename map from disk.  Returns dict of (vin, new_fn) → original_fn."""
    map_path = output_root / _RENAME_MAP_FILE
    if not map_path.exists():
        return {}
    try:
        with open(str(map_path), 'r', encoding='utf-8') as f:
            raw = json.load(f)
        # JSON keys are strings; convert "vin||new_fn" back to (vin, new_fn) tuples
        result = {}
        for key, orig in raw.items():
            parts = key.split("||", 1)
            if len(parts) == 2:
                result[(parts[0], parts[1])] = orig
        print(f"  Rename map: loaded {len(result)} entries from {map_path.name}",
              file=sys.stderr)
        return result
    except Exception as exc:
        print(f"  WARNING: Could not load rename map: {exc}", file=sys.stderr)
        return {}


def save_rename_map(output_root: Path, original_names: dict):
    """Save rename map to disk.  Merges with existing entries (never forgets)."""
    map_path = output_root / _RENAME_MAP_FILE
    # Load existing to merge (don't lose mappings from prior runs)
    existing = {}
    if map_path.exists():
        try:
            with open(str(map_path), 'r', encoding='utf-8') as f:
                existing = json.load(f)
        except Exception:
            pass
    # Merge new entries (new wins on conflict)
    for (vin, new_fn), orig_fn in original_names.items():
        key = f"{vin}||{new_fn}"
        existing[key] = orig_fn
    try:
        with open(str(map_path), 'w', encoding='utf-8') as f:
            json.dump(existing, f, ensure_ascii=False)
        print(f"  Rename map: saved {len(existing)} entries to {map_path.name}",
              file=sys.stderr)
    except Exception as exc:
        print(f"  WARNING: Could not save rename map: {exc}", file=sys.stderr)


def _ocr_cache_lookup(path_str: str) -> tuple:
    """Check persistent cache for a previous OCR result.
    Returns (hit, vins_set, cats_set).  hit=False means must rescan."""
    entry = _ocr_disk_cache.get(path_str)
    if not entry or not entry.get("ocr_used"):
        return (False, set(), set())
    size, mtime = _file_fingerprint(path_str)
    if size != entry.get("size", -1) or mtime != entry.get("mtime", -1):
        return (False, set(), set())  # file changed
    return (True, set(entry.get("vins", [])), set(entry.get("cats", [])))


def _ocr_cache_store(path_str: str, vins: set = None, cats: set = None,
                     reclass_cat: str = "__UNSET__"):
    """Store a scan result in the persistent OCR cache (call from main process).
    Can store VIN scan results, reclass category, or both."""
    size, mtime = _file_fingerprint(path_str)
    entry = _ocr_disk_cache.get(path_str, {})
    entry["size"] = size
    entry["mtime"] = mtime
    entry["ocr_used"] = True
    if vins is not None:
        entry["vins"] = sorted(vins)
    if cats is not None:
        entry["cats"] = sorted(cats)
    if reclass_cat != "__UNSET__":
        entry["reclass_cat"] = reclass_cat  # None or category string
    _ocr_disk_cache[path_str] = entry

# Keywords to detect critical categories from PDF text content (uppercase)
_CONTENT_CAT_KEYWORDS = {
    "Contract Cadru": [
        re.compile(r'CONTRACT\s+CADRU', re.I),
        re.compile(r'CONTRACT\s+DE\s+LEASING', re.I),
        re.compile(r'LEASING\s+OPERA[TȚ]IONAL', re.I),
    ],
    "Subcontract": [
        re.compile(r'SUBCONTRACT', re.I),
    ],
    "CASCO": [
        re.compile(r'CASCO', re.I),
        re.compile(r'FLEXICASCO', re.I),
        re.compile(r'POLI[TȚ][AĂ]\s*DT', re.I),
    ],
    "RCA": [
        re.compile(r'\bRCA\b', re.I),
        re.compile(r'RASPUNDERE\s+CIVIL[AĂ]', re.I),
        re.compile(r'ASIGURARE\s+OBLIGATORIE', re.I),
    ],
}


def _detect_content_categories(text: str) -> set:
    """Detect critical document categories from PDF text content."""
    cats = set()
    for cat, patterns in _CONTENT_CAT_KEYWORDS.items():
        for pat in patterns:
            if pat.search(text):
                cats.add(cat)
                break
    return cats


_OCR_ENABLED = False   # set by main() based on --ocr flag
_OCR_DPI = 150         # render resolution — 150 is plenty for printed text
_OCR_MIN_TEXT = 50     # chars of alphanumeric text below which we try OCR
_OCR_MAX_PAGES = 2     # only OCR first N pages per PDF (VINs are on page 1-2)
_OCR_TESS_CONFIG = '--oem 1 --psm 6'  # LSTM-only, assume uniform text block

# High-accuracy settings for --ocr-rescue
_OCR_RESCUE_DPI = 300
_OCR_RESCUE_MAX_PAGES = 5
_OCR_RESCUE_TESS_CONFIG = '--oem 1 --psm 3'  # fully automatic page segmentation

_ocr_saved_settings = {}

def _ocr_boost_rescue():
    """Temporarily boost OCR settings for high-accuracy _NO_VIN rescue."""
    global _OCR_DPI, _OCR_MAX_PAGES, _OCR_TESS_CONFIG, _ocr_saved_settings
    _ocr_saved_settings = {
        "dpi": _OCR_DPI, "max_pages": _OCR_MAX_PAGES, "config": _OCR_TESS_CONFIG
    }
    _OCR_DPI = _OCR_RESCUE_DPI
    _OCR_MAX_PAGES = _OCR_RESCUE_MAX_PAGES
    _OCR_TESS_CONFIG = _OCR_RESCUE_TESS_CONFIG

def _ocr_restore():
    """Restore OCR settings after rescue phase."""
    global _OCR_DPI, _OCR_MAX_PAGES, _OCR_TESS_CONFIG
    if _ocr_saved_settings:
        _OCR_DPI = _ocr_saved_settings["dpi"]
        _OCR_MAX_PAGES = _ocr_saved_settings["max_pages"]
        _OCR_TESS_CONFIG = _ocr_saved_settings["config"]

def _ocr_pool_init(dpi, max_pages, tess_config):
    """Initializer for ProcessPoolExecutor workers — propagates OCR settings."""
    global _OCR_DPI, _OCR_MAX_PAGES, _OCR_TESS_CONFIG
    _OCR_DPI = dpi
    _OCR_MAX_PAGES = max_pages
    _OCR_TESS_CONFIG = tess_config

def _ocr_page(page) -> str:
    """Render a PyMuPDF page to grayscale image and OCR it with pytesseract."""
    if not HAS_OCR or not HAS_PYMUPDF:
        return ""
    try:
        mat = fitz.Matrix(_OCR_DPI / 72, _OCR_DPI / 72)
        pix = page.get_pixmap(matrix=mat, colorspace=fitz.csGRAY)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        text = pytesseract.image_to_string(img, lang='ron+eng',
                                           config=_OCR_TESS_CONFIG)
        return text
    except Exception:
        return ""


def _needs_ocr(path_str: str) -> bool:
    """Fast pre-filter: open PDF with PyMuPDF, check if first 2 pages
    have sparse text (<50 alnum chars). If all pages have enough text,
    OCR is unnecessary. This avoids submitting text-rich PDFs to the
    slow OCR pool."""
    if not HAS_PYMUPDF:
        return False
    try:
        doc = fitz.open(_long(path_str))
        needs = False
        for i, page in enumerate(doc):
            if i >= _OCR_MAX_PAGES:
                break
            text = page.get_text()
            alnum = sum(1 for c in text if c.isalnum())
            if alnum < _OCR_MIN_TEXT:
                needs = True
                break
        doc.close()
        return needs
    except Exception:
        return False


def _extract_page_text(page, page_num: int, ocr: bool = False) -> str:
    """Get text from a page, falling back to OCR for image-only pages."""
    text = page.get_text()
    # If very little extractable text, try OCR
    if ocr and page_num < _OCR_MAX_PAGES:
        alnum = sum(1 for c in text if c.isalnum())
        if alnum < _OCR_MIN_TEXT:
            ocr_text = _ocr_page(page)
            if ocr_text:
                text = text + "\n" + ocr_text
    return text


def _scan_single_pdf(path_str: str, ocr: bool = False) -> tuple:
    vins = set()
    cats = set()
    # Check persistent OCR cache first (avoids re-OCR on subsequent runs)
    if ocr:
        hit, cached_vins, cached_cats = _ocr_cache_lookup(path_str)
        if hit:
            return (path_str, cached_vins, cached_cats, None)
    try:
        doc = fitz.open(_long(path_str))
        full_text = chr(12).join(
            _extract_page_text(page, i, ocr=ocr) for i, page in enumerate(doc)
        ).upper()
        doc.close()
        vins = {v for v in VIN_PATTERN.findall(full_text) if is_valid_vin(v)}
        cats = _detect_content_categories(full_text)
        return (path_str, vins, cats, None)
    except Exception as e:
        return (path_str, set(), set(), e)


def extract_vins_from_pdf(path: Path) -> set:
    key = str(path)
    if key in _pdf_cache: return _pdf_cache[key]
    if not HAS_PYMUPDF:
        _pdf_cache[key] = set()
        _pdf_content_cats[key] = set()
        return set()
    _, vins, cats, err = _scan_single_pdf(key, ocr=False)  # OCR is post-copy only
    if err: _pdf_stats["failed"] += 1
    else: _pdf_stats["scanned"] += 1; _pdf_stats["vins_found"] += len(vins)
    _pdf_cache[key] = vins
    _pdf_content_cats[key] = cats
    return vins


def _get_partition_dirs(root: Path, range_start: int = 0, range_end: int = 0) -> list:
    all_parts = []
    for d in sorted(root.iterdir()):
        if not d.is_dir(): continue
        if not (d.name.upper().startswith("SINDICALIZARE") or
                d.name.upper().startswith("SINICALIZARE")):
            continue
        all_parts.append(d)
    start = max(0, (range_start or 1) - 1)
    end = range_end if range_end else len(all_parts)
    return all_parts[start:end]


def bulk_prescan_pdfs(root: Path, workers: int, range_start: int = 0,
                      range_end: int = 0):
    """Pre-scan PDFs for VINs using text extraction only (no OCR).
    OCR is reserved for the post-copy reclassification phase."""
    import concurrent.futures
    pdf_paths = []
    seen = set()
    for part_dir in _get_partition_dirs(root, range_start, range_end):
        try:
            for p in part_dir.rglob("*"):
                if p.is_file() and p.suffix.lower() == ".pdf":
                    key = str(p)
                    if key not in seen:
                        seen.add(key)
                        pdf_paths.append(key)
        except PermissionError:
            pass
    if not pdf_paths: return
    print(f"  Pre-scanning {len(pdf_paths)} PDFs with {workers} process(es)...",
          file=sys.stderr, flush=True)
    bar = tqdm(total=len(pdf_paths), desc="PDF pre-scan", unit="pdf",
               bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]")

    def _cb(path_str, vins, cats, err):
        _pdf_cache[path_str] = vins
        _pdf_content_cats[path_str] = cats
        if err: _pdf_stats["failed"] += 1
        else: _pdf_stats["scanned"] += 1; _pdf_stats["vins_found"] += len(vins)
        bar.update(1)

    PDF_TIMEOUT = 120

    if workers <= 1:
        for p in pdf_paths: _cb(*_scan_single_pdf(p, ocr=False))
    else:
        try:
            Executor = concurrent.futures.ProcessPoolExecutor
            with Executor(max_workers=workers) as exe:
                futs = {exe.submit(_scan_single_pdf, p, False): p for p in pdf_paths}
                for f in concurrent.futures.as_completed(futs):
                    p = futs[f]
                    try:
                        _cb(*f.result(timeout=PDF_TIMEOUT))
                    except concurrent.futures.TimeoutError:
                        _cb(p, set(), set(), TimeoutError(f"PDF hung >{PDF_TIMEOUT}s"))
                        tqdm.write(f"  WARNING: PDF scan timed out: {p}")
                    except Exception as exc:
                        _cb(p, set(), set(), exc)
                        tqdm.write(f"  WARNING: PDF scan crashed: {p}: {exc}")
        except Exception as exc:
            tqdm.write(f"  WARNING: Process pool broken ({exc}), "
                       f"falling back to sequential scanning...")
            bar.close()
            bar = tqdm(total=len(pdf_paths), desc="PDF pre-scan (sequential)",
                       unit="pdf",
                       bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} "
                       "[{elapsed}<{remaining}, {rate_fmt}]")
            for p in pdf_paths:
                if p not in _pdf_cache:
                    _cb(*_scan_single_pdf(p, ocr=False))
                else:
                    bar.update(1)
    bar.close()


# ── Collision-safe file helpers ──────────────────────────────────────────────

def _file_hash(path: Path) -> str:
    h = hashlib.md5()
    with open(_long(path), 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''): h.update(chunk)
    return h.hexdigest()


def _files_identical(a: Path, b: Path) -> bool:
    try:
        if a.stat().st_size != b.stat().st_size: return False
        return _file_hash(a) == _file_hash(b)
    except OSError:
        return False


def _safe_dest(src: Path, dst: Path) -> tuple:
    if not _exists(dst): return dst, "ok"
    if _files_identical(src, dst): return dst, "skip"
    stem, suffix, parent = dst.stem, dst.suffix, dst.parent
    m = re.match(r'^(.+?)_(\d+)$', stem)
    base_stem = m.group(1) if m else stem
    for i in range(1, 10000):
        candidate = parent / f"{base_stem}_{i}{suffix}"
        if not _exists(candidate): return candidate, "renamed"
        if _files_identical(src, candidate): return candidate, "skip"
    return dst, "ok"


# ── Change ledger ────────────────────────────────────────────────────────────

@dataclass
class Change:
    action: str       # copy_file, create_folder
    source: str
    destination: str
    reason: str = ""
    parent_folder: str = ""
    vin: str = ""
    status: str = "planned"


class Ledger:
    def __init__(self):
        self.changes: list = []
        self.warnings: list = []
        self.pdf_scans: list = []
        self._planned_dests: dict = {}

    def add(self, action, source, destination, reason="", parent_folder="", vin=""):
        dst_str = str(destination)
        src_str = str(source)
        if action == "copy_file":
            if dst_str in self._planned_dests:
                if self._planned_dests[dst_str] == src_str:
                    return
            self._planned_dests[dst_str] = src_str
        self.changes.append(Change(
            action=action, source=src_str, destination=dst_str,
            reason=reason, parent_folder=parent_folder, vin=vin,
        ))

    def warn(self, msg):
        self.warnings.append(msg)

    def log_pdf_scan(self, path, vins_found):
        self.pdf_scans.append((str(path), sorted(vins_found)))

    def execute(self, dry_run=True, jsonl_path: Path = None, workers: int = 1):
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import threading

        label = "Preview" if dry_run else "Copying"
        bar = tqdm(total=len(self.changes), desc=label, unit="op",
                   bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]")

        jsonl_fh = None
        jsonl_lock = threading.Lock()
        if jsonl_path and not dry_run:
            os.makedirs(str(jsonl_path.parent), exist_ok=True)
            jsonl_fh = open(jsonl_path, 'w', encoding='utf-8')

        def _log_safe(fh, change):
            if fh is None: return
            with jsonl_lock: self._log_change(fh, change)

        def _exec_copy(c):
            src, dst = Path(c.source), Path(c.destination)
            try:
                if not _exists(src):
                    c.status = "skipped"
                    _log_safe(jsonl_fh, c)
                    return
                os.makedirs(_long(dst.parent), exist_ok=True)
                actual, status = _safe_dest(src, dst)
                if status == "skip":
                    c.status = "skipped"
                    _log_safe(jsonl_fh, c)
                    return
                if status == "renamed":
                    c.destination = str(actual)
                # Retry with exponential backoff for WinError 32 (file locked)
                last_err = None
                for attempt in range(_RETRY_ATTEMPTS):
                    try:
                        shutil.copy2(_long(src), _long(actual))
                        c.status = "done"
                        _log_safe(jsonl_fh, c)
                        return
                    except OSError as e:
                        last_err = e
                        if getattr(e, 'winerror', 0) == 32 or 'being used' in str(e):
                            time.sleep(_RETRY_BASE_DELAY * (2 ** attempt))
                            gc.collect()
                        else:
                            raise
                c.status = "failed"
                _log_safe(jsonl_fh, c)
                tqdm.write(f"  ERROR [{c.action}] {src.name}: {last_err}")
            except Exception as e:
                c.status = "failed"
                _log_safe(jsonl_fh, c)
                tqdm.write(f"  ERROR [{c.action}] {src.name}: {e}")

        try:
            i = 0
            total = len(self.changes)
            while i < total:
                c = self.changes[i]

                # Batch consecutive copy_file ops for threaded execution
                if workers > 1 and c.action == "copy_file":
                    batch = []
                    while i < total and self.changes[i].action == "copy_file":
                        batch.append(self.changes[i])
                        i += 1
                    if dry_run:
                        for _ in batch: bar.update(1)
                        continue
                    with ThreadPoolExecutor(max_workers=workers) as pool:
                        futs = {pool.submit(_exec_copy, bc): bc for bc in batch}
                        for f in as_completed(futs):
                            bar.update(1)
                            f.result()
                    continue

                # Sequential: create_folder or single copy_file
                i += 1
                if dry_run:
                    bar.update(1)
                    continue

                if c.action == "create_folder":
                    dst = Path(c.destination)
                    os.makedirs(_long(dst), exist_ok=True)
                    c.status = "done"
                    self._log_change(jsonl_fh, c)
                elif c.action == "copy_file":
                    _exec_copy(c)

                bar.update(1)
        finally:
            bar.close()
            if jsonl_fh: jsonl_fh.close()

    @staticmethod
    def _log_change(fh, c):
        if fh is None: return
        fh.write(json.dumps({
            "action": c.action, "source": c.source, "destination": c.destination,
            "reason": c.reason, "parent_folder": c.parent_folder,
            "vin": c.vin, "status": c.status,
        }, ensure_ascii=False) + "\n")
        fh.flush()

    def write_json(self, path: Path):
        data = {
            "generated": datetime.datetime.now().isoformat(),
            "changes": [
                {"action": c.action, "source": c.source, "destination": c.destination,
                 "reason": c.reason, "parent_folder": c.parent_folder,
                 "vin": c.vin, "status": c.status}
                for c in self.changes
            ],
            "warnings": self.warnings,
        }
        path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8')
        print(f"  Summary JSON:  {path}")


# ── Document categorization for inventory Excel ─────────────────────────────

DOC_CATEGORIES = [
    ("Formular de Livrare (FL)", [
        re.compile(r'^FL\s', re.I),
        re.compile(r'^fl\s', re.I),
        re.compile(r'_FL_Attachment', re.I),
        re.compile(r'_FL\.', re.I),
    ]),
    ("Contract Cadru", [
        re.compile(r'Contract\s+Cadru', re.I),
        re.compile(r'ctr[\s_\.]*cadru', re.I),
        re.compile(r'CTR\.\s*CADRU', re.I),
    ]),
    ("Subcontract", [
        re.compile(r'Subcontract', re.I),
        re.compile(r'_sub\s*\d', re.I),
    ]),
    ("RCA",         [
        re.compile(r'^POLITA_RCA', re.I),
        re.compile(r'^POLITA_', re.I),
    ]),
    ("CASCO",       [
        re.compile(r'CASCO', re.I),
        re.compile(r'FlexiCasco', re.I),
        re.compile(r'Polita\s*DT', re.I),
    ]),
    ("Facturi",     [
        re.compile(r'[Ff]actur[aăi]', re.I),
        re.compile(r'^FF_', re.I),
        re.compile(r'^ff\.pdf$', re.I),
        re.compile(r'^F\.FINALA', re.I),
    ]),
    ("OP Plăți",    [
        re.compile(r'^OP\s', re.I),
    ]),
    ("Cesiune / Supliment", [
        re.compile(r'Cesiune', re.I),
        re.compile(r'Supliment', re.I),
    ]),
    ("TALON / CIV", [
        re.compile(r'TALON', re.I),
        re.compile(r'\bCIV\b', re.I),
        re.compile(r'CIV\+', re.I),
    ]),
    # Configurare/Ofertă and Serie C go to Alte Documente (no dedicated column)
]

# Files to completely ignore (not even Alte Documente)
_IGNORE_FILES = {'desktop.ini', 'Thumbs.db'}

# Priority keywords that ALWAYS win regardless of DOC_CATEGORIES order:
# - Factura: "Factura Cesiune" → Facturi, not Cesiune
# - TALON/CIV: always gets its own column even inside seriec_ files
_FACTURA_PRIORITY = re.compile(r'factur[aăi]', re.I)
_TALON_CIV_PRIORITY = [
    re.compile(r'TALON', re.I),
    re.compile(r'\bCIV\b', re.I),
    re.compile(r'CIV\+', re.I),
]


# Reverse lookup: short filename stem → category
# Built from _CAT_SHORT_NAMES so renamed files are recognized back
_SHORT_NAME_TO_CAT = {}  # populated after _CAT_SHORT_NAMES is defined


def _build_short_name_reverse():
    """Build reverse mapping from short name stems to categories."""
    global _SHORT_NAME_TO_CAT
    for cat, short in _CAT_SHORT_NAMES.items():
        if short:
            _SHORT_NAME_TO_CAT[short.lower()] = cat
    # TALON/CIV variants
    _SHORT_NAME_TO_CAT["talon"] = "TALON / CIV"
    _SHORT_NAME_TO_CAT["civ"] = "TALON / CIV"
    _SHORT_NAME_TO_CAT["talon_civ"] = "TALON / CIV"
    _SHORT_NAME_TO_CAT["talon+civ"] = "TALON / CIV"
    # Backwards compat: old short names from prior runs
    _SHORT_NAME_TO_CAT["supliment_cesiune"] = "Cesiune / Supliment"


def categorize_file(fn: str) -> str:
    # Skip system files entirely
    if fn in _IGNORE_FILES:
        return None
    # ── Recognise our own short names (cc.pdf, subct_1.pdf, etc.) ────
    stem = Path(fn).stem.lower()
    # Strip trailing _N numbering  (e.g. "cc_2" → "cc", "op_14" → "op")
    base = re.sub(r'_\d+$', '', stem)
    cat = _SHORT_NAME_TO_CAT.get(base)
    if cat:
        return cat
    # Factura always wins (highest priority)
    if _FACTURA_PRIORITY.search(fn):
        return "Facturi"
    # TALON / CIV always wins (second priority)
    for pat in _TALON_CIV_PRIORITY:
        if pat.search(fn):
            return "TALON / CIV"
    # Then check all other categories
    for cat_name, patterns in DOC_CATEGORIES:
        for pat in patterns:
            if pat.search(fn):
                return cat_name
    return "Alte Documente"


def build_inventory(output_root: Path, range_start: int = 0, range_end: int = 0,
                    original_names: dict = None) -> dict:
    """Scan output directory: {VIN: {partition, files: {category: [filenames]}}}
    If original_names is provided, uses original filenames for display in Excel.
    original_names maps (vin, renamed_fn) → original_fn.
    Scans ALL subdirectories for VIN folders — output partition names may differ
    from source names (e.g. 'sin_alpha' vs 'SINDICALIZARE ALPHA FINAL')."""
    inventory = {}

    try:
        top_entries = sorted(output_root.iterdir())
    except OSError as exc:
        print(f"  WARNING: Cannot list output directory '{output_root}': {exc}",
              file=sys.stderr)
        return inventory

    part_dirs = [d for d in top_entries if d.is_dir()]
    for part_dir in part_dirs:
        dname = part_dir.name
        # Skip hidden/special dirs and files with extensions
        if dname.startswith(("_", ".")) or "." in dname:
            continue
        try:
            for vin_dir in sorted(part_dir.iterdir()):
                if not vin_dir.is_dir() or not is_vin(vin_dir.name):
                    continue
                vin = vin_dir.name
                if vin not in inventory:
                    inventory[vin] = {
                        "_partition": dname,
                        "_actual_partition": dname,
                        "_files": defaultdict(list),
                    }
                for f in vin_dir.rglob('*'):
                    if f.is_file():
                        cat = categorize_file(f.name)
                        if cat is None:
                            continue
                        rel = f.relative_to(vin_dir)
                        display_name = f.name
                        if original_names:
                            display_name = original_names.get((vin, f.name), f.name)
                        display_rel = str(rel.parent / display_name) if rel.parent != Path('.') \
                            else display_name
                        inventory[vin]["_files"][cat].append(display_rel)
        except PermissionError:
            pass
    return inventory


def build_inventory_from_ledger(ledger, output_root: Path,
                                original_names: dict = None) -> dict:
    """Build inventory purely from the planning ledger.
    The ledger has VIN, partition, and filename data from the source scan.
    No output directory checks needed — the planning step already determined
    what goes where."""
    inventory = {}

    for change in ledger.changes:
        if change.action != "copy_file":
            continue
        vin = change.vin
        if not vin or not is_vin(vin):
            continue
        dest = Path(change.destination)

        # Extract partition name from destination path
        try:
            rel = dest.relative_to(output_root)
        except ValueError:
            continue
        parts = rel.parts
        if len(parts) < 2:
            continue
        partition_name = parts[0]

        if vin not in inventory:
            inventory[vin] = {
                "_partition": partition_name,
                "_actual_partition": partition_name,
                "_files": defaultdict(list),
            }

        # The planned destination filename (after rename if applicable)
        actual_fn = dest.name
        cat = categorize_file(actual_fn)
        if cat is None:
            continue

        # Display name: use original name if available
        display_name = actual_fn
        if original_names:
            display_name = original_names.get((vin, actual_fn), actual_fn)

        # Preserve subdir structure (e.g. contracte/cc.pdf)
        vin_base = output_root / partition_name / vin
        try:
            file_rel = dest.relative_to(vin_base)
        except ValueError:
            file_rel = Path(actual_fn)
        if file_rel.parent != Path('.'):
            display_rel = str(file_rel.parent / display_name)
        else:
            display_rel = display_name

        if display_rel not in inventory[vin]["_files"][cat]:
            inventory[vin]["_files"][cat].append(display_rel)

    return inventory


# ── Category-aware filename renaming ────────────────────────────────────────


def _file_hash(path_str: str) -> str:
    """MD5 hash of file contents for deduplication."""
    h = hashlib.md5()
    try:
        with open(_long(path_str), 'rb') as f:
            for chunk in iter(lambda: f.read(65536), b''):
                h.update(chunk)
    except OSError:
        return f"__error_{path_str}"
    return h.hexdigest()


# Category → short filename (no VIN, no details)
_CAT_SHORT_NAMES = {
    "Contract Cadru":           "cc",
    "Subcontract":              "subct",
    "Cesiune / Supliment":      "ces",
    "Formular de Livrare (FL)": "fl",
    "TALON / CIV":              None,   # special handling
    "CASCO":                    "casco",
    "RCA":                      "rca",
    "OP Plăți":                 "op",
    "Facturi":                  "fact",
}

_build_short_name_reverse()  # populate _SHORT_NAME_TO_CAT


def _detect_talon_civ(fn: str) -> tuple:
    """Returns (has_talon, has_civ) for a filename."""
    has_talon = bool(re.search(r'TALON', fn, re.I))
    has_civ = bool(re.search(r'(?<![A-Za-z])CIV(?![A-Za-z])', fn, re.I))
    return has_talon, has_civ


def _rename_dedup_group(changes, indices, base_name, stats, original_names):
    """Universal dedup+rename: identical files → single {base}.pdf,
    different files → {base}_1.pdf, {base}_2.pdf, etc."""
    if not indices:
        return set()
    remove = set()

    # Hash source files
    hashes = {}
    for idx in indices:
        hashes[idx] = _file_hash(changes[idx].source)

    # Group by hash
    by_hash = defaultdict(list)
    for idx in indices:
        by_hash[hashes[idx]].append(idx)

    unique_hashes = list(by_hash.keys())
    if len(unique_hashes) == 1:
        # All identical — keep first, remove rest
        group = by_hash[unique_hashes[0]]
        keeper = group[0]
        for idx in group[1:]:
            remove.add(idx)
            stats["deduped"] += 1
        c = changes[keeper]
        dst = Path(c.destination)
        new_name = f"{base_name}.pdf"
        original_names[(c.vin, new_name)] = dst.name
        c.destination = str(dst.parent / new_name)
        stats["renamed"] += 1
    else:
        # Multiple different files — dedup within each hash, then number
        counter = 0
        for h in unique_hashes:
            group = by_hash[h]
            keeper = group[0]
            for idx in group[1:]:
                remove.add(idx)
                stats["deduped"] += 1
            counter += 1
            c = changes[keeper]
            dst = Path(c.destination)
            new_name = f"{base_name}_{counter}.pdf"
            original_names[(c.vin, new_name)] = dst.name
            c.destination = str(dst.parent / new_name)
            stats["renamed"] += 1

    return remove


def _rename_talon_civ_group(changes, indices, stats, original_names):
    """TALON/CIV: detect per-file whether it's talon, civ, or both,
    then dedup within each sub-group."""
    if not indices:
        return set()

    # Classify each file
    sub_groups = defaultdict(list)  # base_name -> [indices]
    for idx in indices:
        c = changes[idx]
        fn = Path(c.destination).name
        has_talon, has_civ = _detect_talon_civ(fn)
        if has_talon and has_civ:
            sub_groups["TALON+CIV"].append(idx)
        elif has_talon:
            sub_groups["talon"].append(idx)
        elif has_civ:
            sub_groups["civ"].append(idx)
        else:
            sub_groups["talon_civ"].append(idx)

    remove = set()
    for base, sub_indices in sub_groups.items():
        rm = _rename_dedup_group(changes, sub_indices, base, stats, original_names)
        remove |= rm
    return remove


def plan_category_renames(ledger: Ledger):
    """Category-aware renaming and deduplication of planned copies.
    Returns (stats, original_names) where original_names maps
    (vin, renamed_filename) → original_filename for Excel display."""
    stats = {"renamed": 0, "deduped": 0}
    original_names = {}  # (vin, new_filename) → old_filename

    # Group changes by VIN
    by_vin = defaultdict(list)
    for i, c in enumerate(ledger.changes):
        if c.action != "copy_file":
            continue
        by_vin[c.vin].append(i)

    remove_all = set()

    for vin, indices in by_vin.items():
        # Sub-group by category
        by_cat = defaultdict(list)
        for idx in indices:
            c = ledger.changes[idx]
            fn = Path(c.destination).name
            cat = categorize_file(fn)
            if cat is not None:
                by_cat[cat].append(idx)

        # Apply rename+dedup for each category with a short name
        for cat, short in _CAT_SHORT_NAMES.items():
            if cat not in by_cat:
                continue
            if cat == "TALON / CIV":
                rm = _rename_talon_civ_group(
                    ledger.changes, by_cat[cat], stats, original_names)
            else:
                rm = _rename_dedup_group(
                    ledger.changes, by_cat[cat], short, stats, original_names)
            remove_all |= rm

        # Alte Documente: no rename

    # Remove deduped entries
    if remove_all:
        ledger.changes = [c for i, c in enumerate(ledger.changes) if i not in remove_all]
        ledger._planned_dests = {}
        for c in ledger.changes:
            if c.action == "copy_file":
                ledger._planned_dests[c.destination] = c.source

    return stats, original_names


# ── Post-copy content-based reclassification ─────────────────────────────────
# Patterns to identify document category from PDF *text content*.
# These are broader than filename patterns because the full text is richer.
_CONTENT_CATEGORY_PATTERNS = {
    "Contract Cadru": [
        re.compile(r'Contract\s+Cadru', re.I),
        re.compile(r'Contract\s+de\s+Leasing', re.I),
        re.compile(r'Leasing\s+Opera[tț]ional', re.I),
    ],
    "Subcontract": [
        re.compile(r'Subcontract', re.I),
        re.compile(r'Act\s+Adi[tț]ional', re.I),
    ],
    "CASCO": [
        re.compile(r'\bCASCO\b', re.I),
        re.compile(r'FlexiCasco', re.I),
        re.compile(r'Poli[tț][aă]\s+DT\b', re.I),
    ],
    "RCA": [
        re.compile(r'\bRCA\b'),                        # uppercase only – avoid false positives
        re.compile(r'R[aă]spundere\s+Civil[aă]', re.I),
    ],
    "TALON / CIV": [
        re.compile(r'\bTALON\b', re.I),
        re.compile(r'Certificat\s+de\s+[IÎ]nmatricul', re.I),
        re.compile(r'\bCIV\b'),
    ],
    "Facturi": [
        re.compile(r'FACTUR[AĂ]', re.I),
        re.compile(r'Factur[aă]\s+fiscal[aă]', re.I),
        re.compile(r'Factur[aă]\s+proform[aă]', re.I),
    ],
}

# Priority order for content classification (first match wins)
_CONTENT_PRIORITY = ["Facturi", "TALON / CIV", "Contract Cadru", "Subcontract", "CASCO", "RCA"]

# Test-injectable cache: path_str -> category name (or None)
_reclass_cache: dict = {}


def _scan_pdf_for_category(pdf_path: str, ocr: bool = False) -> Optional[str]:
    """Open a PDF and determine its category from text content.
    Returns the first matching category or None."""
    # Check test/cache first
    if pdf_path in _reclass_cache:
        return _reclass_cache[pdf_path]
    # Check persistent OCR cache for previous reclassification result
    if ocr:
        entry = _ocr_disk_cache.get(pdf_path)
        if entry and entry.get("ocr_used") and "reclass_cat" in entry:
            size, mtime = _file_fingerprint(pdf_path)
            if size == entry.get("size", -1) and mtime == entry.get("mtime", -1):
                val = entry["reclass_cat"]
                return val  # None or category string
    if not HAS_PYMUPDF:
        return None
    try:
        doc = fitz.open(_long(pdf_path))
        text = chr(12).join(
            _extract_page_text(page, i, ocr=ocr) for i, page in enumerate(doc)
        )
        doc.close()
    except Exception:
        return None

    for cat in _CONTENT_PRIORITY:
        for pat in _CONTENT_CATEGORY_PATTERNS[cat]:
            if pat.search(text):
                return cat
    return None


def reclassify_by_content(inventory: dict, output_root: Path, workers: int = 4,
                          ocr: bool = False, rename_on_disk: bool = False):
    """Post-copy phase: scan 'Alte Documente' PDFs by content to find
    miscategorized documents (generic filenames hiding contracts, CASCO, etc.).
    Only scans VINs that have missing critical categories.
    Mutates inventory dict in-place, moving files between categories."""
    if not HAS_PYMUPDF and not _reclass_cache:
        print("  PyMuPDF not installed – skipping content reclassification",
              file=sys.stderr)
        return {"scanned": 0, "reclassified": 0, "vins_checked": 0}

    # Identify which VINs have gaps AND have Alte Documente PDFs to scan
    scan_tasks = []  # (vin, pdf_rel_path, pdf_abs_path)
    critical = set(_CONTENT_PRIORITY)

    for vin, data in inventory.items():
        files = data["_files"]
        # Which critical categories does this VIN already have?
        present = {cat for cat in critical if files.get(cat)}
        missing = critical - present
        if not missing:
            continue
        # Collect Alte Documente PDFs for scanning
        alte = files.get("Alte Documente", [])
        for rel in alte:
            if not rel.lower().endswith('.pdf'):
                continue
            part = data.get("_actual_partition", data["_partition"])
            abs_path = str(output_root / part / vin / rel)
            scan_tasks.append((vin, rel, abs_path, missing))

    if not scan_tasks:
        print("  Content reclassification: no VINs with gaps + Alte Documente PDFs",
              file=sys.stderr)
        return {"scanned": 0, "reclassified": 0, "vins_checked": 0}

    vins_checked = len({t[0] for t in scan_tasks})

    reclassified = 0
    import concurrent.futures
    results = []
    scan_errors = 0
    PDF_TIMEOUT = 30

    # ── Pre-filter: separate cached / text-rich / needs-OCR ──────────
    cached_count = 0
    text_tasks = []
    ocr_scan_tasks = []

    if ocr:
        for t in scan_tasks:
            abs_path = t[2]
            # Check persistent cache first (instant)
            entry = _ocr_disk_cache.get(abs_path)
            if entry and entry.get("ocr_used") and "reclass_cat" in entry:
                size, mtime = _file_fingerprint(abs_path)
                if size == entry.get("size", -1) and mtime == entry.get("mtime", -1):
                    results.append((t[0], t[1], entry["reclass_cat"], t[3]))
                    cached_count += 1
                    continue
            if _needs_ocr(abs_path):
                ocr_scan_tasks.append(t)
            else:
                text_tasks.append(t)
        if cached_count or text_tasks:
            print(f"    Pre-filter: {cached_count} cached, "
                  f"{len(text_tasks)} text-rich (skip OCR), "
                  f"{len(ocr_scan_tasks)} need OCR",
                  file=sys.stderr, flush=True)
    else:
        text_tasks = scan_tasks

    total_to_scan = len(text_tasks) + len(ocr_scan_tasks)
    print(f"  Scanning {total_to_scan} Alte Documente PDFs across "
          f"{vins_checked} VINs for content reclassification...",
          file=sys.stderr, flush=True)

    bar = tqdm(total=total_to_scan, desc="Content scan",
               unit="pdf", bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} "
               "[{elapsed}<{remaining}, {rate_fmt}]")

    def _do_scan_batch(tasks, use_ocr):
        nonlocal scan_errors
        if not tasks:
            return
        timeout = PDF_TIMEOUT if use_ocr else 120
        if workers > 1:
            try:
                with concurrent.futures.ProcessPoolExecutor(
                        max_workers=workers,
                        initializer=_ocr_pool_init,
                        initargs=(_OCR_DPI, _OCR_MAX_PAGES, _OCR_TESS_CONFIG)) as pool:
                    futs = {pool.submit(_scan_pdf_for_category, t[2], use_ocr): t
                            for t in tasks}
                    for f in concurrent.futures.as_completed(futs):
                        bar.update(1)
                        task = futs[f]
                        vin, rel, abs_path, missing = task
                        try:
                            cat = f.result(timeout=timeout)
                            results.append((vin, rel, cat, missing))
                            if use_ocr:
                                _ocr_cache_store(abs_path, reclass_cat=cat)
                        except concurrent.futures.TimeoutError:
                            scan_errors += 1
                            tqdm.write(f"  WARNING: Content scan timed out: {abs_path}")
                        except Exception as exc:
                            scan_errors += 1
                            tqdm.write(f"  WARNING: Content scan failed: {abs_path}: {exc}")
            except Exception:
                for t in tasks:
                    vin, rel, abs_path, missing = t
                    try:
                        cat = _scan_pdf_for_category(abs_path, ocr=use_ocr)
                        results.append((vin, rel, cat, missing))
                        if use_ocr:
                            _ocr_cache_store(abs_path, reclass_cat=cat)
                    except Exception:
                        scan_errors += 1
                    bar.update(1)
        else:
            for t in tasks:
                vin, rel, abs_path, missing = t
                try:
                    cat = _scan_pdf_for_category(abs_path, ocr=use_ocr)
                    results.append((vin, rel, cat, missing))
                    if use_ocr:
                        _ocr_cache_store(abs_path, reclass_cat=cat)
                except Exception:
                    scan_errors += 1
                bar.update(1)

    _do_scan_batch(text_tasks, False)
    _do_scan_batch(ocr_scan_tasks, True)
    bar.close()

    if scan_errors:
        print(f"  Content scan: {scan_errors} PDFs failed (skipped)",
              file=sys.stderr)

    # Apply reclassifications
    for vin, rel, detected_cat, missing in results:
        if detected_cat is None:
            continue
        # Only reclassify if it fills a gap for this VIN
        if detected_cat not in missing:
            continue
        files = inventory[vin]["_files"]
        if rel in files.get("Alte Documente", []):
            new_rel = rel  # default: keep same path

            if rename_on_disk:
                # Rename file on disk to category short name
                short = _CAT_SHORT_NAMES.get(detected_cat)
                if short:
                    data = inventory[vin]
                    part = data.get("_actual_partition", data["_partition"])
                    old_abs = Path(output_root / part / vin / rel)
                    if old_abs.exists():
                        new_name = f"{short}.pdf"
                        new_abs = old_abs.parent / new_name
                        # Handle collision
                        if new_abs.exists() and not _files_identical(old_abs, new_abs):
                            counter = 1
                            while (old_abs.parent / f"{short}_{counter}.pdf").exists():
                                counter += 1
                            new_name = f"{short}_{counter}.pdf"
                            new_abs = old_abs.parent / new_name
                        if new_abs.exists() and _files_identical(old_abs, new_abs):
                            # Duplicate — just remove the old one
                            try:
                                old_abs.unlink()
                            except OSError:
                                pass
                            new_rel = new_name
                        else:
                            try:
                                old_abs.rename(new_abs)
                                new_rel = new_name
                            except OSError as exc:
                                tqdm.write(f"  WARNING: rename failed {rel} → {new_name}: {exc}")

            files["Alte Documente"].remove(rel)
            files[detected_cat].append(new_rel)
            reclassified += 1

    stats = {"scanned": len(scan_tasks), "reclassified": reclassified,
             "vins_checked": vins_checked, "scan_errors": scan_errors}
    if reclassified:
        print(f"  Content reclassification: {reclassified} PDFs moved to correct categories",
              file=sys.stderr)
    return stats


def write_inventory_excel(excel_path: Path, inventory: dict):
    """Write inventory Excel from scratch. No merge with old data."""
    if not HAS_OPENPYXL:
        print("  openpyxl not installed, skipping Excel. pip install openpyxl")
        return

    cat_names = [c[0] for c in DOC_CATEGORIES] + ["Alte Documente"]
    headers = ["VIN", "Partition"] + cat_names + ["Total Files"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(top=thin, bottom=thin, left=thin, right=thin)
    wrap = Alignment(vertical="top", wrap_text=True)

    widths = [20, 38] + [35] * len(cat_names) + [12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = brd
        ws.column_dimensions[c.column_letter].width = w
    ws.freeze_panes = "A2"

    row = 2
    for vin in sorted(inventory.keys()):
        info = inventory[vin]
        partition = info["_partition"]
        files = info["_files"]
        total = sum(len(v) for v in files.values())
        ws.cell(row=row, column=1, value=vin).border = brd
        ws.cell(row=row, column=2, value=partition).border = brd
        for ci, cat in enumerate(cat_names):
            cell = ws.cell(row=row, column=ci + 3,
                           value="\n".join(sorted(files.get(cat, []))))
            cell.border = brd
            cell.alignment = wrap
        ws.cell(row=row, column=len(headers), value=total).border = brd
        row += 1

    from openpyxl.utils import get_column_letter
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{row - 1}"

    wb.save(str(excel_path))
    print(f"  Inventory Excel: {excel_path}  ({row - 2} VINs)")


# ── Folder planning ──────────────────────────────────────────────────────────

def get_parent_vin(folder: Path) -> Optional[str]:
    fl_vins, seriec_vins, other_vins = [], [], []
    try:
        for item in folder.iterdir():
            if not item.is_file() or item.suffix.lower() != '.pdf': continue
            fn = item.name
            m = FL_PATTERN.match(fn)
            if m: fl_vins.append(m.group(1)); continue
            m = SERIEC_PATTERN.match(fn)
            if m: seriec_vins.append(m.group(1)); continue
            m = re.match(r'^([A-Z0-9]{17})[\s_\-]', fn)
            if m and is_valid_vin(m.group(1)): other_vins.append(m.group(1))
    except PermissionError:
        return None
    for pool in (fl_vins, seriec_vins, other_vins):
        if pool: return max(set(pool), key=pool.count)
    return None


def _copy_dir_files(src_dir: Path, dst_dir: Path, ledger: Ledger,
                    parent_folder: str, vin: str, reason: str):
    """Plan copy_file for every file recursively under src_dir → dst_dir."""
    try:
        for item in sorted(src_dir.rglob('*')):
            if item.is_file():
                rel = item.relative_to(src_dir)
                ledger.add("copy_file", item, dst_dir / rel,
                           reason=reason, parent_folder=parent_folder, vin=vin)
    except PermissionError:
        ledger.warn(f"Cannot read '{src_dir}'")


def plan_vin_folder(folder: Path, out_partition: Path, ledger: Ledger):
    """Copy a VIN-named folder's contents to output. Elevate nested VINs."""
    vin = folder.name
    target = out_partition / vin

    try:
        for item in sorted(folder.iterdir()):
            if item.is_dir():
                if is_vin(item.name):
                    _copy_dir_files(item, out_partition / item.name, ledger,
                                    parent_folder=vin, vin=item.name,
                                    reason="Elevate nested VIN")
                else:
                    _copy_dir_files(item, target / item.name, ledger,
                                    parent_folder=vin, vin=vin,
                                    reason="Copy subdir contents")
            elif item.is_file():
                ledger.add("copy_file", item, target / item.name,
                           reason="Copy from VIN folder", parent_folder=vin, vin=vin)
    except PermissionError:
        ledger.warn(f"Cannot read VIN folder '{vin}'")


def plan_multi_car(folder: Path, vin_subdirs: list,
                   out_partition: Path, ledger: Ledger, scan_pdf: bool):
    name = folder.name
    parent_vin = get_parent_vin(folder)
    if not parent_vin:
        # Fall back to the first VIN subdir as parent VIN for loose files
        parent_vin = sorted(vd.name for vd in vin_subdirs)[0]

    vin_subdir_names = {vd.name for vd in vin_subdirs}
    target = out_partition / parent_vin

    # 1. Copy VIN subdirs to output as separate VIN folders
    for vd in vin_subdirs:
        _copy_dir_files(vd, out_partition / vd.name, ledger,
                        parent_folder=name, vin=vd.name,
                        reason="Copy sub-VIN to output")

    # 2. Copy remaining files/subdirs → parent VIN folder in output
    try:
        for item in sorted(folder.iterdir()):
            if item.is_dir():
                if item.name in vin_subdir_names:
                    continue
                _copy_dir_files(item, target / item.name, ledger,
                                parent_folder=name, vin=parent_vin,
                                reason="Copy subdir to parent VIN")
            elif item.is_file():
                ledger.add("copy_file", item, target / item.name,
                           reason="Copy to parent VIN", parent_folder=name, vin=parent_vin)
    except PermissionError:
        ledger.warn(f"Cannot list '{name}' for dissolution")


def plan_flat(folder: Path, out_partition: Path, ledger: Ledger, scan_pdf: bool):
    name = folder.name

    file_fn_vins: dict = {}
    all_fn_vins: set = set()
    all_vins_for_election: set = set()
    try:
        for item in folder.iterdir():
            if not item.is_file(): continue
            fn_vins = set(extract_all_vins(item.name))
            file_fn_vins[item.name] = fn_vins
            all_fn_vins |= fn_vins
            if scan_pdf and item.suffix.lower() == '.pdf':
                pdf_vins = extract_vins_from_pdf(item)
                if pdf_vins:
                    ledger.log_pdf_scan(item, pdf_vins)
                    all_vins_for_election |= pdf_vins
            all_vins_for_election |= fn_vins
    except PermissionError:
        ledger.warn(f"Cannot read '{name}'")
        return

    if not all_fn_vins and not all_vins_for_election:
        # Last resort: check the FOLDER NAME itself for a VIN
        # e.g. "JTEBR3FJ20K323532 - TOYOTA LANDRUISER - PAINEA DE CASA"
        folder_vins = set(extract_all_vins(name))
        if folder_vins:
            all_vins_for_election |= folder_vins
        else:
            # Truly no VINs anywhere — copy to _NO_VIN/original_name/
            if file_fn_vins:  # only if folder actually has files
                no_vin_target = out_partition / "_NO_VIN" / name
                try:
                    for item in sorted(folder.iterdir()):
                        if item.is_dir():
                            _copy_dir_files(item, no_vin_target / item.name, ledger,
                                            parent_folder=name, vin="_NO_VIN",
                                            reason="No VIN found — preserve in _NO_VIN")
                        elif item.is_file():
                            ledger.add("copy_file", item, no_vin_target / item.name,
                                       reason="No VIN found — preserve in _NO_VIN",
                                       parent_folder=name, vin="_NO_VIN")
                except PermissionError:
                    pass
                ledger.warn(f"No VINs in '{name}' ({len(file_fn_vins)} files) → copied to _NO_VIN")
            else:
                ledger.warn(f"No VINs in '{name}' (empty folder)")
            return

    # Pick keeper VIN
    parent_vin = get_parent_vin(folder)
    if parent_vin and parent_vin in (all_fn_vins or all_vins_for_election):
        keeper = parent_vin
    else:
        vin_counts = defaultdict(int)
        for fvins in file_fn_vins.values():
            for v in fvins: vin_counts[v] += 1
        if vin_counts:
            keeper = max(vin_counts, key=vin_counts.get)
        elif all_vins_for_election:
            keeper = list(all_vins_for_election)[0]
        else:
            return

    other_vins = all_fn_vins - {keeper}
    target = out_partition / keeper
    copied_out = set()

    if other_vins:
        for item in folder.iterdir():
            if not item.is_file(): continue
            fn = item.name
            fvins = file_fn_vins.get(fn, set())
            if not fvins or fvins == {keeper}: continue

            if keeper not in fvins:
                primary = sorted(fvins)[0]
                ledger.add("copy_file", item, out_partition / primary / fn,
                           reason="Filename VIN match", parent_folder=name, vin=primary)
                copied_out.add(fn)
                for v in sorted(fvins - {primary}):
                    ledger.add("copy_file", item, out_partition / v / fn,
                               reason="Filename VIN match", parent_folder=name, vin=v)
            else:
                for v in sorted(fvins - {keeper}):
                    ledger.add("copy_file", item, out_partition / v / fn,
                               reason="Filename VIN match", parent_folder=name, vin=v)

    # Copy remaining files → keeper VIN folder in output
    try:
        for item in sorted(folder.iterdir()):
            if item.is_dir():
                if is_vin(item.name):
                    _copy_dir_files(item, out_partition / item.name, ledger,
                                    parent_folder=name, vin=item.name,
                                    reason="Elevate VIN subdir")
                else:
                    _copy_dir_files(item, target / item.name, ledger,
                                    parent_folder=name, vin=keeper,
                                    reason="Copy subdir to keeper VIN")
            elif item.is_file():
                if item.name in copied_out: continue
                ledger.add("copy_file", item, target / item.name,
                           reason="Copy to keeper VIN", parent_folder=name, vin=keeper)
    except PermissionError:
        ledger.warn(f"Cannot list '{name}' for dissolution")


# ── Scanning and planning ───────────────────────────────────────────────────

def scan_and_plan(root: Path, output_root: Path, ledger: Ledger, scan_pdf: bool,
                  range_start: int = 0, range_end: int = 0):
    stats = defaultdict(int)

    all_folders = []
    for part_dir in _get_partition_dirs(root, range_start, range_end):
        try:
            for cdir in sorted(part_dir.iterdir()):
                if cdir.is_dir():
                    all_folders.append((cdir, part_dir.name))
        except PermissionError:
            pass

    bar = tqdm(all_folders, desc="Scanning folders", unit="folder",
               bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]")

    for cdir, partition_name in bar:
        bar.set_postfix_str(cdir.name[:40], refresh=False)
        out_partition = output_root / merge_partition_name(partition_name)

        if is_vin(cdir.name):
            stats["vin_named"] += 1
            plan_vin_folder(cdir, out_partition, ledger)
            continue

        vin_subdirs = []
        has_files = False
        has_other_dirs = False
        try:
            for sub in cdir.iterdir():
                if sub.is_dir():
                    if is_vin(sub.name): vin_subdirs.append(sub)
                    else: has_other_dirs = True
                elif sub.is_file(): has_files = True
        except PermissionError:
            ledger.warn(f"Cannot read '{cdir.name}'")
            stats["error"] += 1
            continue

        if not vin_subdirs and not has_files and not has_other_dirs:
            continue

        if vin_subdirs:
            stats["multi_car"] += 1
            plan_multi_car(cdir, vin_subdirs, out_partition, ledger, scan_pdf)
        else:
            stats["flat"] += 1
            plan_flat(cdir, out_partition, ledger, scan_pdf)

    return stats


def plan_pdf_cross_copies(ledger: Ledger, output_root: Path):
    """Post-planning pass: for every PDF being copied, check its text content
    for VINs and cross-copy to all matching VIN folders.
    PDFs with >MAX_CROSS_COPY_VINS are skipped to avoid bloat."""
    stats = {"cross_copied": 0, "skipped_too_many": 0, "pdfs_checked": 0}

    # Build VIN → output partition path from planned changes
    vin_partition: dict = {}  # vin -> partition dir path
    for c in ledger.changes:
        if not c.vin or c.action not in ("copy_file", "create_folder"):
            continue
        dst = Path(c.destination)
        # destination looks like: output_root / partition / VIN / file.pdf
        # or output_root / partition / VIN / subdir / file.pdf
        # We need the partition component
        try:
            rel = dst.relative_to(output_root)
            parts = rel.parts
            if len(parts) >= 2:
                partition_name = parts[0]
                vin_partition[c.vin] = output_root / partition_name
        except ValueError:
            pass

    # Track what's already planned: (source_str, vin) pairs
    already_planned = set()
    for c in ledger.changes:
        if c.action == "copy_file":
            already_planned.add((c.source, c.vin))

    # Process all planned copy_file ops for PDFs
    # Snapshot the list since we'll append
    original_changes = list(ledger.changes)
    for c in original_changes:
        if c.action != "copy_file":
            continue
        src = Path(c.source)
        if src.suffix.lower() != '.pdf':
            continue

        # Get content VINs from cache
        content_vins = _pdf_cache.get(str(src), set())
        if not content_vins:
            continue
        stats["pdfs_checked"] += 1

        if len(content_vins) > MAX_CROSS_COPY_VINS:
            stats["skipped_too_many"] += 1
            ledger.warn(f"PDF '{src.name}' has {len(content_vins)} VINs in content, "
                        f"skipping cross-copy (limit={MAX_CROSS_COPY_VINS})")
            continue

        # Cross-copy to every VIN folder mentioned in PDF content
        for vin in sorted(content_vins):
            if (str(src), vin) in already_planned:
                continue
            if vin not in vin_partition:
                # VIN not seen in any planned changes — skip
                # (could be a VIN from a different partition range not being processed)
                continue
            out_part = vin_partition[vin]
            dest = out_part / vin / src.name
            ledger.add("copy_file", src, dest,
                       reason="PDF content VIN cross-copy",
                       parent_folder=c.parent_folder, vin=vin)
            already_planned.add((str(src), vin))
            stats["cross_copied"] += 1

    return stats


# Categories that every VIN folder should ideally have
_CRITICAL_CATEGORIES = {"Contract Cadru", "Subcontract", "CASCO", "RCA"}

# Patterns to identify a PDF's critical category from filename (for gap-fill matching)
_CONTRACT_PATTERNS = [
    re.compile(r'Contract\s+Cadru', re.I),
    re.compile(r'ctr[\s_\.]*cadru', re.I),
    re.compile(r'CTR\.\s*CADRU', re.I),
    re.compile(r'Contract\s+de\s+Leasing', re.I),
    re.compile(r'LO\s+Contract', re.I),
]
_SUBCONTRACT_PATTERNS = [
    re.compile(r'Subcontract', re.I),
    re.compile(r'_sub\s*\d', re.I),
    re.compile(r'^VIEW_Subcontract', re.I),
]
_CASCO_PATTERNS = [
    re.compile(r'CASCO', re.I),
    re.compile(r'FlexiCasco', re.I),
    re.compile(r'Polita\s*DT', re.I),
]
_RCA_PATTERNS = [
    re.compile(r'POLITA_RCA', re.I),
    re.compile(r'^POLITA_', re.I),
    re.compile(r'\bRCA\b', re.I),
]


def _pdf_critical_category(fn: str) -> Optional[str]:
    """If filename indicates a critical category, return it."""
    # Factura takes priority - skip if it's a factura
    if _FACTURA_PRIORITY.search(fn):
        return None
    for pat in _SUBCONTRACT_PATTERNS:
        if pat.search(fn):
            return "Subcontract"
    for pat in _CONTRACT_PATTERNS:
        if pat.search(fn):
            return "Contract Cadru"
    for pat in _CASCO_PATTERNS:
        if pat.search(fn):
            return "CASCO"
    for pat in _RCA_PATTERNS:
        if pat.search(fn):
            return "RCA"
    return None


def plan_contract_gap_fill(ledger: Ledger, output_root: Path):
    """Last sweep: find VINs missing critical documents (contracts, subcontracts,
    CASCO, RCA) and fill gaps by cross-copying from PDFs that mention those VINs
    in their content, even if those PDFs exceed the normal cross-copy VIN limit.
    Uses BOTH filename patterns AND PDF text content to identify categories."""
    stats = {"gap_filled": 0, "vins_with_gaps": 0}

    # 1. Build VIN → set of critical categories already planned
    vin_categories: dict = defaultdict(set)  # vin -> set of critical cats
    vin_partition: dict = {}  # vin -> partition output path
    already_planned: set = set()  # (source_str, vin)

    for c in ledger.changes:
        if c.action != "copy_file":
            continue
        already_planned.add((c.source, c.vin))
        dst = Path(c.destination)
        fn = dst.name
        # Check filename for critical category
        fn_cat = _pdf_critical_category(fn)
        if fn_cat and c.vin:
            vin_categories[c.vin].add(fn_cat)
        # Also check PDF content categories (from pre-scan)
        if c.vin:
            content_cats = _pdf_content_cats.get(c.source, set())
            for cc in content_cats:
                if cc in _CRITICAL_CATEGORIES:
                    vin_categories[c.vin].add(cc)
        # Track partitions
        if c.vin:
            try:
                rel = dst.relative_to(output_root)
                parts = rel.parts
                if len(parts) >= 2:
                    vin_partition[c.vin] = output_root / parts[0]
            except ValueError:
                pass

    # 2. Find VINs with gaps
    vins_needing = defaultdict(set)  # vin -> set of missing critical cats
    for vin in vin_partition:
        missing = _CRITICAL_CATEGORIES - vin_categories.get(vin, set())
        if missing:
            vins_needing[vin] = missing
    if not vins_needing:
        return stats
    stats["vins_with_gaps"] = len(vins_needing)

    # 3. Build reverse index: for each PDF source, what critical categories + what VINs
    # Uses BOTH filename patterns and PDF text content keywords
    pdf_info: dict = {}  # source_str -> (set of critical cats, content_vins)
    for c in ledger.changes:
        if c.action != "copy_file":
            continue
        src_str = c.source
        if src_str in pdf_info:
            continue
        src = Path(src_str)
        if src.suffix.lower() != '.pdf':
            continue
        content_vins = _pdf_cache.get(src_str, set())
        if not content_vins:
            continue
        # Combine filename category + content categories
        cats = set()
        fn_cat = _pdf_critical_category(src.name)
        if fn_cat:
            cats.add(fn_cat)
        content_cats = _pdf_content_cats.get(src_str, set())
        cats |= (content_cats & _CRITICAL_CATEGORIES)
        if cats:
            pdf_info[src_str] = (cats, content_vins)

    # 4. For each VIN with gaps, find PDFs that can fill them
    for vin, missing_cats in vins_needing.items():
        if vin not in vin_partition:
            continue
        out_part = vin_partition[vin]
        for src_str, (cats, content_vins) in pdf_info.items():
            matching = cats & missing_cats
            if not matching:
                continue
            if vin not in content_vins:
                continue
            if (src_str, vin) in already_planned:
                continue
            src = Path(src_str)
            dest = out_part / vin / src.name
            filled_cat = sorted(matching)[0]  # pick one for reason label
            ledger.add("copy_file", src, dest,
                       reason=f"Gap-fill: {filled_cat} from PDF content",
                       parent_folder="", vin=vin)
            already_planned.add((src_str, vin))
            for cat in matching:
                vin_categories[vin].add(cat)
                missing_cats.discard(cat)
            stats["gap_filled"] += 1
            if not missing_cats:
                break

    return stats



# ── Rescan: fix existing output in-place ─────────────────────────────────────


def _scan_pdf_full(pdf_path: str, ocr: bool = False) -> tuple:
    """Scan a single PDF, return (path, vins, cats, error). Wrapper for pool."""
    if not HAS_PYMUPDF:
        return (pdf_path, set(), set(), None)
    return _scan_single_pdf(pdf_path, ocr=ocr)


def _short_name_for_category(cat: str, fn: str) -> str:
    """Return the short filename for a detected category. Handles TALON/CIV."""
    if cat == "TALON / CIV":
        has_talon, has_civ = _detect_talon_civ(fn)
        if has_talon and has_civ:
            return "TALON+CIV.pdf"
        elif has_talon:
            return "talon.pdf"
        elif has_civ:
            return "civ.pdf"
        return "talon_civ.pdf"
    short = _CAT_SHORT_NAMES.get(cat)
    if short:
        return f"{short}.pdf"
    return None  # Alte Documente or unknown — keep original name


def _place_file_with_short_name(src: Path, target_dir: Path, cat: str):
    """Move/copy src into target_dir using category short name, with dedup."""
    short_fn = _short_name_for_category(cat, src.name)
    if short_fn:
        dst = target_dir / short_fn
    else:
        dst = target_dir / src.name

    dst.parent.mkdir(parents=True, exist_ok=True)
    if dst.exists():
        if _files_identical(src, dst):
            try:
                src.unlink()
            except OSError:
                pass
            return dst
        # Collision: number it
        stem, suffix = dst.stem, dst.suffix
        counter = 1
        while (dst.parent / f"{stem}_{counter}{suffix}").exists():
            counter += 1
        dst = dst.parent / f"{stem}_{counter}{suffix}"
    shutil.move(str(src), str(dst))
    return dst


def rescan_rescue_no_vin(output_root: Path, workers: int = 4, ocr: bool = False):
    """Scan _NO_VIN folders: OCR PDFs for VINs + categories, move to VIN
    folders with proper short names.  Files with no VIN stay in _NO_VIN."""
    import concurrent.futures

    moved_files = 0
    rescued_folders = 0
    # (abs_pdf_path, no_vin_folder_abs, partition_dir_abs)
    scan_tasks = []

    for part_dir in sorted(output_root.iterdir()):
        if not part_dir.is_dir():
            continue
        no_vin_dir = part_dir / "_NO_VIN"
        if not no_vin_dir.exists():
            continue
        for folder in sorted(no_vin_dir.iterdir()):
            if not folder.is_dir():
                continue
            for f in folder.rglob("*"):
                if f.is_file() and f.suffix.lower() == ".pdf":
                    scan_tasks.append((str(f), str(folder), str(part_dir)))

    if not scan_tasks:
        print("  No _NO_VIN PDFs to rescan.", file=sys.stderr)
        return {"moved": 0, "rescued_folders": 0}

    ocr_label = " +OCR" if ocr else ""
    print(f"  Scanning {len(scan_tasks)} _NO_VIN PDFs for VINs & categories{ocr_label}...",
          file=sys.stderr, flush=True)

    # Per-file results: pdf_path -> (vins, cats)
    pdf_results = {}
    # Per-folder aggregated VINs
    folder_vins = defaultdict(set)
    PDF_TIMEOUT = 30

    # ── Pre-filter: separate cached / text-rich / needs-OCR ──────────
    cached_tasks = []   # already in OCR cache → instant
    text_tasks = []     # text-rich → fast text-only scan
    ocr_tasks = []      # sparse text → needs actual OCR

    if ocr:
        for task in scan_tasks:
            pdf_path = task[0]
            hit, cached_vins, cached_cats = _ocr_cache_lookup(pdf_path)
            if hit:
                cached_tasks.append(task)
                pdf_results[pdf_path] = (cached_vins, cached_cats)
                folder_vins[task[1]] |= cached_vins
            elif _needs_ocr(pdf_path):
                ocr_tasks.append(task)
            else:
                text_tasks.append(task)
        if cached_tasks or text_tasks:
            print(f"    Pre-filter: {len(cached_tasks)} cached, "
                  f"{len(text_tasks)} text-rich (skip OCR), "
                  f"{len(ocr_tasks)} need OCR",
                  file=sys.stderr, flush=True)
    else:
        text_tasks = scan_tasks

    total = len(text_tasks) + len(ocr_tasks)
    bar = tqdm(total=total, desc="NO_VIN rescue",
               unit="pdf", bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} "
               "[{elapsed}<{remaining}, {rate_fmt}]")

    # ── Phase A: fast text-only scan for text-rich PDFs ──────────────
    def _collect(pdf_path, folder, vins, cats, err, use_ocr):
        pdf_results[pdf_path] = (vins, cats)
        folder_vins[folder] |= vins
        if use_ocr and not err:
            _ocr_cache_store(pdf_path, vins=vins, cats=cats)
        bar.update(1)

    if text_tasks:
        if workers > 1:
            try:
                with concurrent.futures.ProcessPoolExecutor(
                        max_workers=workers,
                        initializer=_ocr_pool_init,
                        initargs=(_OCR_DPI, _OCR_MAX_PAGES, _OCR_TESS_CONFIG)) as pool:
                    futs = {pool.submit(_scan_pdf_full, t[0], False): t
                            for t in text_tasks}
                    for f in concurrent.futures.as_completed(futs):
                        t = futs[f]
                        try:
                            _, vins, cats, err = f.result(timeout=120)
                            _collect(t[0], t[1], vins, cats, err, False)
                        except Exception as exc:
                            bar.update(1)
                            tqdm.write(f"  WARNING: Scan failed: {t[0]}: {exc}")
            except Exception:
                for t in text_tasks:
                    if t[0] not in pdf_results:
                        try:
                            _, vins, cats, err = _scan_pdf_full(t[0], False)
                            _collect(t[0], t[1], vins, cats, err, False)
                        except Exception:
                            bar.update(1)
        else:
            for t in text_tasks:
                try:
                    _, vins, cats, err = _scan_pdf_full(t[0], False)
                    _collect(t[0], t[1], vins, cats, err, False)
                except Exception:
                    bar.update(1)

    # ── Phase B: OCR scan for sparse-text PDFs ───────────────────────
    if ocr_tasks:
        if workers > 1:
            try:
                with concurrent.futures.ProcessPoolExecutor(
                        max_workers=workers,
                        initializer=_ocr_pool_init,
                        initargs=(_OCR_DPI, _OCR_MAX_PAGES, _OCR_TESS_CONFIG)) as pool:
                    futs = {pool.submit(_scan_pdf_full, t[0], True): t
                            for t in ocr_tasks}
                    for f in concurrent.futures.as_completed(futs):
                        t = futs[f]
                        try:
                            _, vins, cats, err = f.result(timeout=PDF_TIMEOUT)
                            _collect(t[0], t[1], vins, cats, err, True)
                        except Exception as exc:
                            bar.update(1)
                            tqdm.write(f"  WARNING: OCR failed: {t[0]}: {exc}")
            except Exception:
                for t in ocr_tasks:
                    if t[0] not in pdf_results:
                        try:
                            _, vins, cats, err = _scan_pdf_full(t[0], True)
                            _collect(t[0], t[1], vins, cats, err, True)
                        except Exception:
                            bar.update(1)
        else:
            for t in ocr_tasks:
                try:
                    _, vins, cats, err = _scan_pdf_full(t[0], True)
                    _collect(t[0], t[1], vins, cats, err, True)
                except Exception:
                    bar.update(1)

    bar.close()

    # Move files from _NO_VIN folders to VIN folders
    processed_folders = set()
    for _, folder_abs, part_abs in scan_tasks:
        if folder_abs in processed_folders:
            continue
        processed_folders.add(folder_abs)

        vins = folder_vins.get(folder_abs, set())
        if not vins:
            # Fallback: check folder name for VIN
            folder_name = Path(folder_abs).name
            vins = set(extract_all_vins(folder_name))

        if not vins:
            continue

        primary_vin = sorted(vins)[0]
        part_dir = Path(part_abs)
        target = part_dir / primary_vin
        folder = Path(folder_abs)

        try:
            target.mkdir(parents=True, exist_ok=True)
            for item in sorted(folder.rglob("*")):
                if not item.is_file():
                    continue
                # Determine category: first try filename, then OCR-detected
                cat = categorize_file(item.name)
                pdf_key = str(item)
                if cat == "Alte Documente" and pdf_key in pdf_results:
                    _, ocr_cats = pdf_results[pdf_key]
                    # Pick highest-priority detected category
                    for prio_cat in _CONTENT_PRIORITY:
                        if prio_cat in ocr_cats:
                            cat = prio_cat
                            break

                _place_file_with_short_name(item, target, cat)
                moved_files += 1

            # Cross-copy to other VINs
            for other_vin in sorted(vins - {primary_vin}):
                other_target = part_dir / other_vin
                other_target.mkdir(parents=True, exist_ok=True)
                for item in target.rglob("*"):
                    if not item.is_file():
                        continue
                    dst = other_target / item.name
                    if not dst.exists():
                        dst.parent.mkdir(parents=True, exist_ok=True)
                        shutil.copy2(str(item), str(dst))

            # Clean up empty _NO_VIN/folder
            try:
                remaining = list(folder.rglob("*"))
                if not any(r.is_file() for r in remaining):
                    shutil.rmtree(str(folder), ignore_errors=True)
                    rescued_folders += 1
            except Exception:
                pass
        except Exception as exc:
            tqdm.write(f"  WARNING: Failed to move {folder.name}: {exc}")

    # Clean up empty _NO_VIN dirs
    for part_dir in sorted(output_root.iterdir()):
        if not part_dir.is_dir():
            continue
        no_vin_dir = part_dir / "_NO_VIN"
        if no_vin_dir.exists():
            try:
                remaining = list(no_vin_dir.rglob("*"))
                if not any(r.is_file() for r in remaining):
                    shutil.rmtree(str(no_vin_dir), ignore_errors=True)
            except Exception:
                pass

    stats = {"moved": moved_files, "rescued_folders": rescued_folders}
    if moved_files:
        print(f"  _NO_VIN rescue: moved {moved_files} files, "
              f"rescued {rescued_folders} folders", file=sys.stderr)
    else:
        print(f"  _NO_VIN rescue: no VINs discovered in any _NO_VIN PDFs",
              file=sys.stderr)
    return stats


def rescan_apply_renames(output_root: Path):
    """Phase 2: Walk VIN folders and apply category renames to files on disk.
    Only renames files that still have their original (long) names."""
    stats = {"renamed": 0, "deduped": 0}
    original_names = {}  # for Excel

    for part_dir in sorted(output_root.iterdir()):
        if not part_dir.is_dir():
            continue
        for vin_dir in sorted(part_dir.iterdir()):
            if not vin_dir.is_dir():
                continue
            if vin_dir.name.startswith("_"):
                continue

            # Group files by category
            by_cat = defaultdict(list)  # cat -> [(path, filename)]
            for f in vin_dir.rglob("*"):
                if not f.is_file():
                    continue
                cat = categorize_file(f.name)
                if cat is not None:
                    by_cat[cat].append(f)

            vin = vin_dir.name

            for cat, short in _CAT_SHORT_NAMES.items():
                files = by_cat.get(cat, [])
                if not files:
                    continue

                if cat == "TALON / CIV":
                    # Sub-classify
                    sub_groups = defaultdict(list)
                    for f in files:
                        has_talon, has_civ = _detect_talon_civ(f.name)
                        if has_talon and has_civ:
                            sub_groups["TALON+CIV"].append(f)
                        elif has_talon:
                            sub_groups["talon"].append(f)
                        elif has_civ:
                            sub_groups["civ"].append(f)
                        else:
                            sub_groups["talon_civ"].append(f)
                    for base, sub_files in sub_groups.items():
                        _rescan_rename_group(sub_files, base, vin, stats, original_names)
                else:
                    _rescan_rename_group(files, short, vin, stats, original_names)

    if stats["renamed"]:
        print(f"  Renames applied: {stats['renamed']} files renamed, "
              f"{stats['deduped']} duplicates removed", file=sys.stderr)
    else:
        print(f"  Renames: all files already have short names", file=sys.stderr)
    return stats, original_names


def _rescan_rename_group(files: list, base_name: str, vin: str,
                         stats: dict, original_names: dict):
    """Rename a group of files on disk to {base_name}.pdf with dedup."""
    if not files:
        return

    # Check if already renamed (skip if so)
    expected_names = {f"{base_name}.pdf"} | {f"{base_name}_{i}.pdf" for i in range(1, 50)}
    all_already_renamed = all(f.name in expected_names for f in files)
    if all_already_renamed:
        return

    # Hash for dedup
    hashes = {}
    for f in files:
        try:
            h = hashlib.md5()
            with open(_long(f), 'rb') as fh:
                for chunk in iter(lambda: fh.read(65536), b''):
                    h.update(chunk)
            hashes[f] = h.hexdigest()
        except OSError:
            hashes[f] = f"__err_{f}"

    by_hash = defaultdict(list)
    for f in files:
        by_hash[hashes[f]].append(f)

    unique_hashes = list(by_hash.keys())

    if len(unique_hashes) == 1:
        # All identical — keep one, remove rest
        group = by_hash[unique_hashes[0]]
        keeper = group[0]
        for f in group[1:]:
            try:
                f.unlink()
                stats["deduped"] += 1
            except OSError:
                pass
        # Rename keeper
        new_name = f"{base_name}.pdf"
        new_path = keeper.parent / new_name
        if keeper.name != new_name:
            if new_path.exists() and _files_identical(keeper, new_path):
                keeper.unlink()
                stats["deduped"] += 1
            elif not new_path.exists():
                original_names[(vin, new_name)] = keeper.name
                keeper.rename(new_path)
                stats["renamed"] += 1
    else:
        # Multiple different — dedup within hash groups, then number
        counter = 0
        for h in unique_hashes:
            group = by_hash[h]
            keeper = group[0]
            for f in group[1:]:
                try:
                    f.unlink()
                    stats["deduped"] += 1
                except OSError:
                    pass
            counter += 1
            new_name = f"{base_name}_{counter}.pdf"
            new_path = keeper.parent / new_name
            if keeper.name != new_name and not new_path.exists():
                original_names[(vin, new_name)] = keeper.name
                keeper.rename(new_path)
                stats["renamed"] += 1


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    import argparse
    global _OCR_ENABLED

    parser = argparse.ArgumentParser(
        description="Reorganize SIN vehicle folders by VIN (copy to output directory)")
    parser.add_argument("--execute", action="store_true",
                        help="Apply changes (default: dry run)")
    parser.add_argument("--root", type=str, default=str(SIN_ROOT),
                        help="Source root (default: C:\\SIN)")
    parser.add_argument("--output", type=str, default=str(OUTPUT_ROOT),
                        help="Output root (default: C:\\SIN_Changed)")
    parser.add_argument("--no-pdf", action="store_true",
                        help="Skip PDF content scanning")
    parser.add_argument("--workers", type=int, default=None,
                        help="Parallel threads (default: CPU count, capped at 8)")
    parser.add_argument("--range-start", type=int, default=0,
                        help="Process from partition index, 1-based (0=beginning)")
    parser.add_argument("--range-end", type=int, default=0,
                        help="Process up to partition index, inclusive (0=end)")
    parser.add_argument("--rename-files", action="store_true",
                        help="Standardize PDF filenames in output")
    parser.add_argument("--excel", type=str, default=None,
                        help="Path to centralized inventory Excel (default: <output>/inventory.xlsx)")
    parser.add_argument("--no-content-scan", action="store_true",
                        help="Skip content-based reclassification of Alte Documente PDFs")
    parser.add_argument("--inventory-only", action="store_true",
                        help="Skip all copying; just rebuild inventory Excel from existing output")
    parser.add_argument("--ocr", action="store_true",
                        help="Enable OCR for image-only PDFs (requires pytesseract + Tesseract)")
    parser.add_argument("--ocr-rescue", action="store_true",
                        help="Enable high-accuracy OCR only for _NO_VIN rescue scans "
                             "(DPI 300, 5 pages, skip OCR for content reclassification). "
                             "Implies --ocr for rescue phase only.")
    parser.add_argument("--rescan", action="store_true",
                        help="Rescan existing output: rescue _NO_VIN folders via OCR, "
                             "re-apply renames, rebuild Excel. Use with --inventory-only.")
    args = parser.parse_args()

    root = Path(args.root)
    output_root = Path(args.output)

    # ── Inventory-only / Rescan mode ────────────────────────────────────────
    if args.inventory_only:
        if not output_root.exists():
            print(f"ERROR: Output '{output_root}' does not exist.", file=sys.stderr)
            sys.exit(1)
        if not root.exists():
            print(f"ERROR: Source '{root}' does not exist.", file=sys.stderr)
            sys.exit(1)

        workers = args.workers or min(8, os.cpu_count() or 4)
        ocr_rescue_only = args.ocr_rescue and HAS_OCR
        _OCR_ENABLED = (args.ocr or args.ocr_rescue) and HAS_OCR
        if (args.ocr or args.ocr_rescue) and not HAS_OCR:
            print("pytesseract not installed. pip install pytesseract Pillow", file=sys.stderr)
        # --ocr-rescue: OCR only for rescue phase, not reclassification
        ocr_for_rescue = _OCR_ENABLED
        ocr_for_reclass = args.ocr and HAS_OCR and not ocr_rescue_only
        do_rescan = args.rescan
        excel_path = Path(args.excel) if args.excel else output_root / "inventory.xlsx"

        scan_pdf = not args.no_pdf
        if scan_pdf and not HAS_PYMUPDF:
            scan_pdf = False

        rs, re_ = args.range_start, args.range_end
        all_partitions = _get_partition_dirs(root)
        selected = _get_partition_dirs(root, rs, re_)

        mode_label = "RESCAN + INVENTORY" if do_rescan else "INVENTORY ONLY"
        print(f"{'='*70}", file=sys.stderr)
        print(f"SIN Folder Reorganizer – {mode_label} (from source)", file=sys.stderr)
        print(f"  Source:     {root}", file=sys.stderr)
        print(f"  Output:     {output_root}", file=sys.stderr)
        print(f"  Excel:      {excel_path}", file=sys.stderr)
        print(f"  PDF scan:   {'ON' if scan_pdf else 'OFF'}"
              + (f"  (workers={workers})" if scan_pdf else ""), file=sys.stderr)
        if ocr_rescue_only:
            print(f"  OCR:        RESCUE ONLY (DPI {_OCR_RESCUE_DPI}, "
                  f"{_OCR_RESCUE_MAX_PAGES} pages, workers={workers})", file=sys.stderr)
        elif _OCR_ENABLED:
            print(f"  OCR:        ON (pytesseract, first {_OCR_MAX_PAGES} pages, "
                  f"workers={workers})", file=sys.stderr)
        print(f"  Partitions: {len(selected)}/{len(all_partitions)}"
              + (f"  (range {rs or 1}–{re_ or len(all_partitions)})" if rs or re_ else " (all)"),
              file=sys.stderr)
        if args.rename_files:
            print(f"  Rename:     ON (standardize PDF filenames)", file=sys.stderr)
        if do_rescan:
            print(f"  Rescan:     ON (rescue _NO_VIN, reclassify+rename on disk)",
                  file=sys.stderr)
        for p in selected:
            print(f"    • {p.name}", file=sys.stderr)
        print(f"{'='*70}\n", file=sys.stderr)

        # ── Phase 1: Scan source to get original filenames ────────────────
        if scan_pdf:
            bulk_prescan_pdfs(root, workers, rs, re_)

        ledger = Ledger()
        stats = scan_and_plan(root, output_root, ledger, scan_pdf, rs, re_)

        if scan_pdf:
            plan_pdf_cross_copies(ledger, output_root)
            plan_contract_gap_fill(ledger, output_root)

        # Category renames → gives us original_names mapping
        original_names = {}
        if args.rename_files:
            _, original_names = plan_category_renames(ledger)

        # Save rename map for future runs
        if original_names:
            save_rename_map(output_root, original_names)

        # ── Phase 2: Rescan operations (optional) ────────────────────────
        if do_rescan:
            if ocr_for_rescue:
                load_ocr_cache(output_root)
                if ocr_rescue_only:
                    _ocr_boost_rescue()

            rescue_stats = rescan_rescue_no_vin(
                output_root, workers=workers, ocr=ocr_for_rescue)

            if ocr_rescue_only:
                _ocr_restore()

            # Build inventory from actual output for content reclassification
            # (needs real disk paths, not ledger's planned paths)
            print(f"  Building inventory for content reclassification...")
            tmp_inv = build_inventory(output_root)

            if not args.no_content_scan and scan_pdf:
                print(f"  Checking Alte Documente PDFs for miscategorized documents...")
                reclass_stats = reclassify_by_content(
                    tmp_inv, output_root, workers, ocr=ocr_for_reclass,
                    rename_on_disk=True)
                if reclass_stats.get("reclassified", 0):
                    print(f"  Reclassified: {reclass_stats['reclassified']} PDFs "
                          f"(scanned {reclass_stats['scanned']} across "
                          f"{reclass_stats['vins_checked']} VINs) — renamed on disk")

            # Apply short-name renames on disk; merge with source original_names
            rescan_rename_stats, rescan_orig = rescan_apply_renames(output_root)
            # Merge: source-planned names take priority, rescan fills gaps
            for k, v in rescan_orig.items():
                if k not in original_names:
                    original_names[k] = v
            save_rename_map(output_root, original_names)

            if ocr_for_rescue:
                save_ocr_cache(output_root)

        # ── Phase 3: Build final inventory + write Excel ─────────────────
        # Build purely from ledger — it has original filenames + partition names
        rename_map = load_rename_map(output_root)
        all_orig = rename_map or {}
        if original_names:
            all_orig.update(original_names)

        print(f"  Building inventory from planning ledger...", file=sys.stderr)
        inventory = build_inventory_from_ledger(
            ledger, output_root, original_names=all_orig or None)
        print(f"  Inventory: {len(inventory)} VINs", file=sys.stderr)

        write_inventory_excel(excel_path, inventory)
        print(f"\n  Done. Excel written to {excel_path}", file=sys.stderr)
        return

    if not root.exists():
        print(f"ERROR: Source '{root}' does not exist.", file=sys.stderr)
        sys.exit(1)

    scan_pdf = not args.no_pdf
    if scan_pdf and not HAS_PYMUPDF:
        print("PyMuPDF not installed. pip install pymupdf", file=sys.stderr)
        print("Falling back to filename-only VIN detection.\n", file=sys.stderr)
        scan_pdf = False

    _OCR_ENABLED = (args.ocr or args.ocr_rescue) and HAS_OCR
    ocr_rescue_only = args.ocr_rescue and HAS_OCR
    if (args.ocr or args.ocr_rescue) and not HAS_OCR:
        print("pytesseract not installed. pip install pytesseract Pillow", file=sys.stderr)
        print("OCR disabled.\n", file=sys.stderr)

    workers = args.workers or min(8, os.cpu_count() or 4)
    rs, re_ = args.range_start, args.range_end

    all_partitions = _get_partition_dirs(root)
    selected = _get_partition_dirs(root, rs, re_)

    mode = "EXECUTE" if args.execute else "DRY RUN"
    print(f"{'='*70}", file=sys.stderr)
    print(f"SIN Folder Reorganizer v3 (copy mode)", file=sys.stderr)
    print(f"  Source:     {root}", file=sys.stderr)
    print(f"  Output:     {output_root}", file=sys.stderr)
    print(f"  Mode:       {mode}", file=sys.stderr)
    print(f"  PDF scan:   {'ON' if scan_pdf else 'OFF'}"
          + (f"  (workers={workers})" if scan_pdf else ""), file=sys.stderr)
    if ocr_rescue_only:
        print(f"  OCR:        RESCUE ONLY (no effect in --execute, use with --rescan)",
              file=sys.stderr)
    elif _OCR_ENABLED:
        print(f"  OCR:        ON (pytesseract, post-copy phases only, "
              f"first {_OCR_MAX_PAGES} pages, 30s timeout)",
              file=sys.stderr)
    print(f"  Execution:  {'threaded' if workers > 1 else 'sequential'}"
          + (f"  (workers={workers})" if workers > 1 else ""), file=sys.stderr)
    print(f"  Partitions: {len(selected)}/{len(all_partitions)}"
          + (f"  (range {rs or 1}–{re_ or len(all_partitions)})" if rs or re_ else " (all)"),
          file=sys.stderr)
    if args.rename_files:
        print(f"  Rename:     ON (standardize PDF filenames)", file=sys.stderr)
    if args.no_content_scan:
        print(f"  Content scan: OFF (--no-content-scan)", file=sys.stderr)
    for p in selected:
        print(f"    • {p.name}", file=sys.stderr)
    print(f"{'='*70}\n", file=sys.stderr)

    # Phase 1: PDF pre-scan
    if scan_pdf:
        bulk_prescan_pdfs(root, workers, rs, re_)

    # Phase 2: Plan copies
    ledger = Ledger()
    stats = scan_and_plan(root, output_root, ledger, scan_pdf, rs, re_)

    # Phase 2.5: PDF content cross-copy
    cross_stats = {}
    if scan_pdf:
        cross_stats = plan_pdf_cross_copies(ledger, output_root)

    # Phase 2.75: Contract/subcontract gap-fill sweep
    gap_stats = {}
    if scan_pdf:
        gap_stats = plan_contract_gap_fill(ledger, output_root)

    # Phase 3: Category-aware renaming + deduplication
    rename_stats = {}
    original_names = {}  # (vin, renamed_fn) → original_fn
    if args.rename_files:
        rename_stats, original_names = plan_category_renames(ledger)

    # Summary
    print(f"\n{'='*70}")
    print(f"SUMMARY")
    print(f"{'='*70}")
    print(f"  Already VIN-named:   {stats.get('vin_named', 0)}")
    print(f"  Multi-car (subdirs): {stats.get('multi_car', 0)}")
    print(f"  Flat descriptive:    {stats.get('flat', 0)}")
    print(f"  Errors:              {stats.get('error', 0)}")
    print(f"  Total copies:        {len(ledger.changes)}")
    print(f"  Warnings:            {len(ledger.warnings)}")
    if scan_pdf:
        print(f"  PDFs scanned:        {_pdf_stats['scanned']}")
        print(f"  PDFs failed:         {_pdf_stats['failed']}")
        print(f"  VINs from PDFs:      {_pdf_stats['vins_found']}")
    if cross_stats:
        print(f"  PDF cross-copies:    {cross_stats.get('cross_copied', 0)}")
        if cross_stats.get('skipped_too_many', 0):
            print(f"  PDFs skipped (>{MAX_CROSS_COPY_VINS} VINs): "
                  f"{cross_stats['skipped_too_many']}")
    if gap_stats:
        print(f"  Contract gap-fills:  {gap_stats.get('gap_filled', 0)}"
              f"  ({gap_stats.get('vins_with_gaps', 0)} VINs had gaps)")
    if rename_stats:
        print(f"  Filenames renamed:   {rename_stats.get('renamed', 0)} (in output)")
        if rename_stats.get('deduped', 0):
            print(f"  Duplicates removed:  {rename_stats['deduped']}")

    if ledger.warnings:
        print(f"\nWARNINGS (first 20):")
        for w in ledger.warnings[:20]:
            print(f"  - {w}")
        if len(ledger.warnings) > 20:
            print(f"  ... and {len(ledger.warnings) - 20} more")

    # Execute
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    jsonl_path = None
    if args.execute and ledger.changes:
        os.makedirs(str(output_root), exist_ok=True)
        jsonl_path = output_root / f"log_{ts}.jsonl"
        print(f"\n  Streaming log: {jsonl_path}")

    if ledger.changes:
        ledger.execute(dry_run=not args.execute, jsonl_path=jsonl_path, workers=workers)

    # Reports
    print(f"\n{'='*70}")
    print(f"OUTPUT")
    print(f"{'='*70}")

    if args.execute:
        json_path = output_root / f"log_{ts}.json"
        ledger.write_json(json_path)

        done = sum(1 for c in ledger.changes if c.status == "done")
        failed = sum(1 for c in ledger.changes if c.status == "failed")
        skipped = sum(1 for c in ledger.changes if c.status == "skipped")
        print(f"\nCopy complete: {done} done, {skipped} skipped (identical), {failed} failed")

        # Build and write inventory Excel
        excel_path = Path(args.excel) if args.excel else output_root / "inventory.xlsx"
        print(f"\n  Building inventory...")
        # Primary: from ledger (has original names, guaranteed correct)
        inventory = build_inventory_from_ledger(
            ledger, output_root, original_names=original_names or None)
        # Supplement: directory scan catches any extras
        dir_inv = build_inventory(output_root, original_names=original_names or None)
        for vin, data in dir_inv.items():
            if vin not in inventory:
                inventory[vin] = data
            else:
                for cat, files in data["_files"].items():
                    existing = set(inventory[vin]["_files"].get(cat, []))
                    for f in files:
                        if f not in existing:
                            inventory[vin]["_files"][cat].append(f)

        # Persist rename map for future --inventory-only runs
        if original_names:
            save_rename_map(output_root, original_names)

        # Phase 5: Content-based reclassification of Alte Documente
        # --ocr-rescue: skip OCR here (only for _NO_VIN rescue)
        ocr_for_reclass = _OCR_ENABLED and not ocr_rescue_only
        reclass_stats = {}
        if scan_pdf and not args.no_content_scan:
            if ocr_for_reclass:
                load_ocr_cache(output_root)
            print(f"  Checking Alte Documente PDFs for miscategorized documents...")
            reclass_stats = reclassify_by_content(
                inventory, output_root, workers, ocr=ocr_for_reclass)
            if reclass_stats.get("reclassified", 0):
                print(f"  Reclassified: {reclass_stats['reclassified']} PDFs "
                      f"(scanned {reclass_stats['scanned']} across "
                      f"{reclass_stats['vins_checked']} VINs)")
            if ocr_for_reclass:
                save_ocr_cache(output_root)

        write_inventory_excel(excel_path, inventory)
    else:
        print(f"\nDRY RUN complete. Use --execute to copy files.")


if __name__ == "__main__":
    main()